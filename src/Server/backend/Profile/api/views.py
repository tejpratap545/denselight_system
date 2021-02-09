from ...Appraisals.models import peerAppraisal, User_Appraisal_List
from ..permissions import IsOwner
from .serializers import *
from backend.GnC.models import Departments
from backend.Profile.models import ResetPasswordToken
from backend.Profile.permissions import IsHrManager
from django.core.cache import cache
from django.db.models import Count, Q
from django.http import HttpResponse
from rest_framework import generics, status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.generics import get_object_or_404
from rest_framework.pagination import PageNumberPagination
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.viewsets import ModelViewSet

import json


class ProfileInfoView(generics.RetrieveAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = ProfileInfoSerializer

    def get_object(self):
        return self.request.user.profile


class DepartmentViewSet(ModelViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class = DepartmentSerializer
    queryset = Departments.objects.all()


class ListEmployees(generics.ListAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = EmployeeSerializer
    queryset = (
        Profile.objects.prefetch_related(
            "first_Reporting_Manager",
            "department",
            "first_Reporting_Manager__department",
            "second_Reporting_Manager",
            "second_Reporting_Manager__department",
        )
        .only(
            "id",
            "name",
            "email",
            "typeOfEmployee",
            "second_Reporting_Manager",
            "first_Reporting_Manager",
            "gender",
            "department",
            "job_Title",
            "date_Of_Hire",
            "resign_date",
        )
        .filter(user__is_active=True)
    )


class ReginListEmployees(generics.ListAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = EmployeeSerializer
    queryset = (
        Profile.objects.prefetch_related(
            "first_Reporting_Manager",
            "department",
            "first_Reporting_Manager__department",
            "second_Reporting_Manager",
            "second_Reporting_Manager__department",
        )
        .only(
            "id",
            "name",
            "email",
            "typeOfEmployee",
            "second_Reporting_Manager",
            "first_Reporting_Manager",
            "gender",
            "department",
            "job_Title",
            "date_Of_Hire",
            "resign_date",
        )
        .filter(user__is_active=False)
    )


class ShortListEmployees(generics.ListAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = ShortEmployeeSerializer
    queryset = Profile.objects.only(
        "id",
        "name",
        "email",
    ).filter(user__is_active=True)


class CreateProfile(generics.CreateAPIView):
    permission_classes = [IsHrManager]
    serializer_class = ProfileCreateSerializer
    queryset = Profile.objects.all()


class ProfileView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsHrManager]

    queryset = Profile.objects.prefetch_related(
        "user",
        "department",
        "first_Reporting_Manager",
        "second_Reporting_Manager",
        "first_Reporting_Manager__department",
        "second_Reporting_Manager__department",
    )

    def get_serializer_class(self):
        if self.request.method == "GET":
            return ProfileInfoSerializer
        return ProfileSerializer


class ChangePassword(generics.CreateAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = ChangePasswordSerializer


class SetPassword(generics.CreateAPIView):
    permission_classes = [IsHrManager]
    serializer_class = SetPasswordSerializer


class NotificationViewSet(ModelViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class = NotificationSerializer
    pagination_class = PageNumberPagination

    def get_queryset(self):
        return (
            Notification.objects.filter(user=self.request.user.profile)
            .order_by("-created_at")
            .annotate(unseen=Count("seen", filter=Q(seen=False)))
        )


class StatusView(generics.RetrieveAPIView):
    serializer_class = StatusSerializer
    permission_classes = [IsAuthenticated]

    def get_object(self):
        return {
            "a1": User_Appraisal_List.objects.filter(
                employee=self.request.user.profile, overall_appraisal__status="Stage 1"
            ).count(),
            "a2": User_Appraisal_List.objects.filter(
                employee=self.request.user.profile, overall_appraisal__status="Stage 1B"
            ).count(),
            "a3": User_Appraisal_List.objects.filter(
                employee=self.request.user.profile, overall_appraisal__status="Stage 2"
            ).count(),
            "a4": peerAppraisal.objects.filter(
                manager=self.request.user.profile, completion="Uncompleted"
            ).count(),
        }


@api_view(["POST"])
@permission_classes([IsHrManager])
def change_role(request):
    try:

        User.objects.filter(profile__id=request.data.get("id")).update(
            role=request.data.get("role")
        )
        return Response("Success")
    except:
        return Response("Errors", status=status.HTTP_400_BAD_REQUEST)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def set_profile_picture(request):
    try:
        request.user.profile.profile_Picture = request.FILES["profilePicture"]
        request.user.profile.save()
        return Response("Success")
    except:
        return Response("Errors", status=status.HTTP_400_BAD_REQUEST)


@api_view(["POST"])
def get_supervisor(request):
    profile = get_object_or_404(Profile, user__username=request.data.get("username"))
    return Response({"Supervisor": profile.first_Reporting_Manager.name})


class GuideView(ModelViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class = GuideSerializer
    queryset = Guide.objects.all()


from django.conf import settings
from django.core.mail import send_mail

import secrets


@api_view(["POST"])
def get_token(request):
    email = request.data.get("email", "")
    if email != "" or email is not None:
        token = secrets.token_hex(3)
        user = get_object_or_404(User, email=email)

        ResetPasswordToken.objects.create(user=user, token=token)

        send_mail(
            "Password reset token",
            f"Hi {user.username} your one time token for reset password is {token} ",
            settings.OFFICIAL_MAIL,
            [user.email],
        )

        return Response({"msg": "Sucessfully send token "})

    return Response({"msg": "Errors"}, status=status.HTTP_400_BAD_REQUEST)


@api_view(["POST"])
def reset_password(request):
    token = request.data.get("token", "")
    password1 = request.data.get("password1", "")
    password2 = request.data.get("password2", "")
    email = request.data.get("email", "")
    if (
        token != ""
        and password1 != ""
        and password2 != ""
        and password1 == password2
        and email != ""
    ):
        try:
            user = get_object_or_404(User, email=email)
            ResetPasswordToken.objects.filter(token=token, user=user).delete()
            user.set_password(password1)
            user.save()
            return Response("Success")
        except:
            return Response({"msg": "Errors"}, status=status.HTTP_400_BAD_REQUEST)

    return Response({"msg": "Errors"}, status=status.HTTP_400_BAD_REQUEST)


from django.utils import timezone


@api_view(["POST"])
@permission_classes([IsHrManager])
def resign_employee(request):
    try:
        id = request.data.get("id")
        profile = get_object_or_404(Profile, id=id)
        profile.resign_date = timezone.now().date()
        profile.user.is_active = False
        profile.save()
        profile.user.save()
        profile.user_appraisal_list_set.update(is_closed=True)
        return Response("User is successfully resignd")

    except:
        return Response("An error accured", status=status.HTTP_400_BAD_REQUEST)


@api_view(["POST"])
@permission_classes([IsHrManager])
def check_username(request):
    try:
        username = request.data.get("username")
        if not User.objects.filter(username=username).exists():
            return Response("username is available")
        return Response("username not is available", status=status.HTTP_400_BAD_REQUEST)
    except:
        return Response("username not is available", status=status.HTTP_400_BAD_REQUEST)


@api_view(["POST"])
def check_password(request):
    password = request.data.get("password")
    if check_password_strength(password):
        return HttpResponse("Password is good")
    return HttpResponse("Weak Password", status=status.HTTP_400_BAD_REQUEST)


@api_view(["POST"])
@permission_classes([IsHrManager])
def check_email(request):
    try:
        email = request.data.get("email")
        if not User.objects.filter(email=email).exists():
            return Response("email is available")
        return Response("email not is available", status=status.HTTP_400_BAD_REQUEST)
    except:
        return Response("email not is available", status=status.HTTP_400_BAD_REQUEST)


from backend.Profile.backend import check_password_strength
from io import BytesIO
from openpyxl import load_workbook

import datetime


@api_view(["POST"])
@permission_classes([IsHrManager])
def bulk_profile_upload(request):
    password = request.data.get("password", "DenselightPassword1234")
    f = request.FILES["file"].read()
    if not f:
        return Response("No file ", status=status.HTTP_400_BAD_REQUEST)
    wb = load_workbook(filename=BytesIO(f))
    result = []
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        username = row[0].value
        email = row[1].value
        name = row[2].value

        role = row[3].value
        job_title = row[4].value
        nric = row[5].value
        address = row[6].value
        division_center = row[7].value
        department = row[8].value
        if not name or name == "":
            continue
        return_user = {"user": name, "errors": []}
        result.append(return_user)

        if not username or username == "":
            return_user["errors"].append("User Must have username")
            continue

        if not email or email == "":
            return_user["errors"].append("User Must have Email")
            continue

        try:
            if not User.objects.filter(username=username).exists():
                return_user["errors"].append("Username must be unique.")
                continue
        except:
            return_user["errors"].append("Not valid username field")

        try:
            if not User.objects.filter(email=email).exists():
                return_user["errors"].append("Email must be unique.")
                continue
        except:
            return_user["errors"].append("Not valid email field")
        if not department or department == "":
            return_user["errors"].append("Deaprtment must be include")
            continue

        try:

            if not Departments.objects.filter(
                Q(name=department) | Q(id=department)
            ).exists():
                continue
        except:
            return_user["errors"].append("Deaprtment must be valid")

        if role not in ["Admin", "HRManager", "Hr", "Manager", "Admin", "Employee"]:
            return_user["errors"].append("Employee role muxt be valid.")
            continue
        user = User.objects.create_user(
            username=username, email=email, password=password, role=role
        )
        Profile.objects.create(
            user=user,
            name=name,
            email=email,
            job_title=job_title,
            nric=nric,
            address=address,
            division_center=division_center,
            department=department,
        )

    for row in ws.iter_rows(min_row=2):
        username = row[0].value
        email = row[1].value
        name = row[2].value
        first_reporting_manager_value = row[9].value
        second_reporting_manager_value = row[10].value
        if not Profile.objects.filter(
            Q(id=first_reporting_manager_value)
            | Q(user__email=first_reporting_manager_value)
            | Q(name=first_reporting_manager_value)
            | Q(user__username=first_reporting_manager_value)
        ).exists():
            for i in result:
                if i.user == name:
                    i.errors.append("Can Not set first reporting manager")

            continue

        if not Profile.objects.filter(
            Q(id=second_reporting_manager_value)
            | Q(user__email=second_reporting_manager_value)
            | Q(name=second_reporting_manager_value)
            | Q(user__username=second_reporting_manager_value)
        ).exists():
            for i in result:
                if i.user == name:
                    i.errors.append("Can Not set second reporting manager")

            continue

        first_reporting_manager = Profile.objects.filter(
            Q(id=first_reporting_manager_value)
            | Q(user__email=first_reporting_manager_value)
            | Q(name=first_reporting_manager_value)
            | Q(user__username=first_reporting_manager_value)
        )
        second_reporting_manager = Profile.objects.filter(
            Q(id=second_reporting_manager_value)
            | Q(user__email=second_reporting_manager_value)
            | Q(name=second_reporting_manager_value)
            | Q(user__username=second_reporting_manager_value)
        )

        Profile.objects.filter(email=email).update(
            first_Reporting_Manager=first_reporting_manager,
            second_reporting_manager=second_reporting_manager,
        )

    return Response(result)
